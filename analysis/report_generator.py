"""
Excel Report Generator
Generates comprehensive security analysis reports in Excel format
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime
import os
import config

class ReportGenerator:
    """Generates Excel reports from analysis data"""
    
    def __init__(self, analysis_results, logs_dataframe, output_dir=None):
        """
        Initialize report generator
        
        Args:
            analysis_results: Dictionary of analysis results
            logs_dataframe: Pandas DataFrame of logs
            output_dir: Directory to save report
        """
        self.analysis_results = analysis_results
        self.df = logs_dataframe
        self.output_dir = output_dir or config.REPORTS_PATH
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
        # Define styles
        self.header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.subheader_fill = PatternFill(start_color="38BDF8", end_color="38BDF8", fill_type="solid")
        self.subheader_font = Font(bold=True, color="FFFFFF", size=11)
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate(self):
        """Generate complete report"""
        try:
            os.makedirs(self.output_dir, exist_ok=True)
            
            self._create_summary_sheet()
            self._create_logs_sheet()
            self._create_top_attackers_sheet()
            self._create_analytics_sheet()
            self._create_risk_sheet()
            self._create_statistics_sheet()
            
            filename = f"Security_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            self.wb.save(filepath)
            return filepath, filename
        except Exception as e:
            print(f"Error generating report: {str(e)}")
            return None, None
    
    def _create_summary_sheet(self):
        """Create summary sheet"""
        ws = self.wb.create_sheet("Summary", 0)
        
        # Title
        ws['A1'] = "CYBER THREAT ANALYSIS REPORT"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        ws.merge_cells('A1:D1')
        ws['A1'].alignment = self.center_alignment
        
        # Date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws.merge_cells('A2:D2')
        
        # Summary statistics
        row = 4
        ws[f'A{row}'] = "SUMMARY STATISTICS"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        stats = self.analysis_results.get('statistics', {})
        
        summary_data = [
            ("Total Logs", stats.get('total_logs', 0)),
            ("Total Attacks", stats.get('total_attacks', 0)),
            ("Failed Logins", stats.get('failed_logins', 0)),
            ("Unique IPs", stats.get('unique_ips', 0)),
            ("Unique Users", stats.get('unique_users', 0)),
            ("Unique Countries", stats.get('unique_countries', 0)),
            ("Unique Attack Types", stats.get('unique_attack_types', 0)),
            ("Critical Threats", stats.get('critical_threats', 0)),
            ("High Risk IPs", stats.get('high_risk_ips', 0)),
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'B{row}'].alignment = self.center_alignment
            row += 1
        
        # Success vs Failed
        row += 1
        ws[f'A{row}'] = "SUCCESS VS FAILED LOGINS"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ratio = self.analysis_results.get('success_vs_failed_ratio', {})
        ws[f'A{row}'] = "Successful Logins"
        ws[f'B{row}'] = ratio.get('success', 0)
        row += 1
        ws[f'A{row}'] = "Failed Logins"
        ws[f'B{row}'] = ratio.get('failed', 0)
        
        # Auto-fit columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _create_logs_sheet(self):
        """Create logs sheet"""
        ws = self.wb.create_sheet("Logs")
        
        if not self.df.empty:
            for r_idx, row in enumerate(dataframe_to_rows(self.df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = self.header_font
                        cell.fill = self.header_fill
                        cell.alignment = self.center_alignment
                    cell.border = self.border
            
            # Auto-fit columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_top_attackers_sheet(self):
        """Create top attackers sheet"""
        ws = self.wb.create_sheet("Top Attackers")
        
        # Title
        ws['A1'] = "TOP ATTACKING IPs"
        ws['A1'].font = self.subheader_font
        ws['A1'].fill = self.subheader_fill
        ws.merge_cells('A1:B1')
        
        # Headers
        ws['A2'] = "IP Address"
        ws['B2'] = "Attack Count"
        for cell in [ws['A2'], ws['B2']]:
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
        
        # Data
        top_ips = self.analysis_results.get('top_attacking_ips', {})
        row = 3
        for ip, count in top_ips.items():
            ws[f'A{row}'] = ip
            ws[f'B{row}'] = count
            ws[f'B{row}'].alignment = self.center_alignment
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_analytics_sheet(self):
        """Create analytics sheet"""
        ws = self.wb.create_sheet("Analytics")
        
        row = 1
        
        # Attack Type Distribution
        ws[f'A{row}'] = "ATTACK TYPE DISTRIBUTION"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Attack Type"
        ws[f'B{row}'] = "Count"
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        row += 1
        attack_dist = self.analysis_results.get('attack_type_distribution', {})
        for attack_type, count in attack_dist.items():
            ws[f'A{row}'] = attack_type
            ws[f'B{row}'] = count
            row += 1
        
        row += 2
        
        # Top Countries
        ws[f'A{row}'] = "TOP ATTACKING COUNTRIES"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Country"
        ws[f'B{row}'] = "Count"
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        row += 1
        top_countries = self.analysis_results.get('top_countries', {})
        for country, count in top_countries.items():
            ws[f'A{row}'] = country
            ws[f'B{row}'] = count
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _create_risk_sheet(self):
        """Create risk analysis sheet"""
        ws = self.wb.create_sheet("Risk Analysis")
        
        # Title
        ws['A1'] = "IP RISK ASSESSMENT"
        ws['A1'].font = self.subheader_font
        ws['A1'].fill = self.subheader_fill
        ws.merge_cells('A1:D1')
        
        # Headers
        headers = ['IP Address', 'Risk Score', 'Risk Level', 'Reasons']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
        
        # Data
        risk_analysis = self.analysis_results.get('risk_analysis', {})
        row = 3
        for ip, risk_data in risk_analysis.items():
            ws[f'A{row}'] = ip
            ws[f'B{row}'] = risk_data.get('score', 0)
            ws[f'C{row}'] = risk_data.get('level', 'Unknown')
            reasons = ', '.join(risk_data.get('reasons', []))
            ws[f'D{row}'] = reasons
            
            # Color code risk level
            level = risk_data.get('level', '')
            if level == 'Critical':
                ws[f'C{row}'].fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
            elif level == 'High':
                ws[f'C{row}'].fill = PatternFill(start_color="F97316", end_color="F97316", fill_type="solid")
            elif level == 'Medium':
                ws[f'C{row}'].fill = PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid")
            else:
                ws[f'C{row}'].fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
            
            ws[f'C{row}'].font = Font(bold=True, color="FFFFFF")
            row += 1
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 50
    
    def _create_statistics_sheet(self):
        """Create statistics sheet"""
        ws = self.wb.create_sheet("Statistics")
        
        row = 1
        
        # Protocol Usage
        ws[f'A{row}'] = "PROTOCOL USAGE"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "Protocol"
        ws[f'B{row}'] = "Count"
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        row += 1
        protocol_usage = self.analysis_results.get('protocol_usage', {})
        for protocol, count in protocol_usage.items():
            ws[f'A{row}'] = protocol
            ws[f'B{row}'] = count
            row += 1
        
        row += 2
        
        # Brute Force Attempts
        ws[f'A{row}'] = "BRUTE FORCE DETECTION"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "IP Address"
        ws[f'B{row}'] = "Failed Attempts"
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        row += 1
        brute_force = self.analysis_results.get('brute_force_attempts', {})
        for ip, count in brute_force.items():
            ws[f'A{row}'] = ip
            ws[f'B{row}'] = count
            row += 1
        
        row += 2
        
        # Port Scans
        ws[f'A{row}'] = "PORT SCANNING DETECTION"
        ws[f'A{row}'].font = self.subheader_font
        ws[f'A{row}'].fill = self.subheader_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        ws[f'A{row}'] = "IP Address"
        ws[f'B{row}'] = "Unique Ports"
        for cell in [ws[f'A{row}'], ws[f'B{row}']]:
            cell.font = self.header_font
            cell.fill = self.header_fill
        
        row += 1
        port_scans = self.analysis_results.get('port_scans', {})
        for ip, count in port_scans.items():
            ws[f'A{row}'] = ip
            ws[f'B{row}'] = count
            row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
